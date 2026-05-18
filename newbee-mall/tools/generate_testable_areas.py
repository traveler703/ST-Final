# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT = Path(__file__).resolve().parents[1] / "NewBeeMall_Testable_Areas.xlsx"

HEADERS = [
    "ID",
    "系统区域",
    "模块",
    "可测试功能点",
    "入口页面/API",
    "测试对象类型",
    "前置条件/角色",
    "重点输入/状态",
    "建议测试技术",
    "典型测试点",
    "风险优先级",
    "适合自动化",
    "备注",
]

ROWS = [
    ["TA-001", "前台商城", "首页", "首页加载与商品推荐展示", "GET /, /index, /index.html", "页面/集成", "游客", "轮播图、分类、新品、热门、推荐商品", "等价类; UI检查; 数据一致性", "首页模块是否显示; 空数据时页面是否正常; 点击商品/分类是否跳转", "中", "是", "依赖首页配置、分类和商品数据"],
    ["TA-002", "前台商城", "首页", "轮播图跳转", "首页轮播图", "UI/导航", "游客", "redirectUrl、图片 URL、排序", "等价类; 链接有效性", "有效链接跳转; ## 或空链接不跳转; 图片加载失败处理", "中", "是", "可从后台轮播图配置联动测试"],
    ["TA-003", "前台商城", "商品分类", "三级分类导航展示", "首页分类导航", "页面/数据", "游客", "一级/二级/三级分类、排序、删除标记", "等价类; 边界值", "分类少于/等于/超过首页展示上限; 无子分类; 已删除分类不显示", "中", "是", "Constants.INDEX_CATEGORY_NUMBER=10"],
    ["TA-004", "前台商城", "商品搜索", "关键词搜索商品", "GET /search?keyword=...", "页面/API", "游客", "keyword、page、goodsCategoryId、orderBy", "等价类; 边界值; 安全测试", "空关键词; 正常关键词; 不存在关键词; 特殊字符; SQL注入字符串; 分页边界", "高", "是", "商品发现入口，适合重点测"],
    ["TA-005", "前台商城", "商品搜索", "分类筛选搜索", "GET /search?goodsCategoryId=...", "页面/API", "游客", "categoryId、page", "等价类; 边界值", "有效三级分类; 一级/二级分类; 不存在分类; 已删除分类", "中", "是", "检查分类路径与结果一致性"],
    ["TA-006", "前台商城", "商品搜索", "搜索分页", "GET /search?page=...", "页面/API", "游客", "page、limit=10", "边界值", "page=1; 最后一页; 超出最大页; 0; 负数; 非数字", "中", "是", "Constants.GOODS_SEARCH_PAGE_LIMIT=10"],
    ["TA-007", "前台商城", "商品详情", "商品详情展示", "GET /goods/detail/{goodsId}", "页面/API", "游客", "goodsId、商品状态、库存", "等价类; 边界值", "有效上架商品; 不存在ID; 下架商品; 库存为0; 长商品名/富文本", "高", "是", "后续购物车依赖该入口"],
    ["TA-008", "前台商城", "会员注册", "注册页面展示", "GET /register", "页面", "游客", "页面静态资源、验证码", "UI检查; 可用性测试", "表单显示; 验证码加载; 移动端布局; 链接跳转登录", "中", "是", "注册流程前置"],
    ["TA-009", "前台商城", "会员注册", "用户注册成功", "POST /register", "API/业务流程", "游客", "loginName、password、verifyCode", "等价类; 边界值; 决策表", "有效用户名密码验证码; 注册后能登录; 密码是否MD5存储", "高", "是", "需先获取 /common/mall/kaptcha session"],
    ["TA-010", "前台商城", "会员注册", "注册输入校验", "POST /register", "API", "游客", "空用户名、空密码、空验证码、错误验证码", "决策表; 等价类", "任一字段为空; 验证码错误; 验证码过期; 重复用户名", "高", "是", "ServiceResultEnum 有明确错误分支"],
    ["TA-011", "前台商城", "会员注册", "重复用户名处理", "POST /register", "API/数据库", "游客", "loginName 已存在", "等价类", "重复注册返回 SAME_LOGIN_NAME_EXIST; 数据库不新增重复用户", "高", "是", "涉及唯一业务规则"],
    ["TA-012", "前台商城", "会员登录", "登录页面展示", "GET /login", "页面", "游客", "验证码、表单、跳转", "UI检查; 可用性测试", "页面可访问; 验证码刷新; 注册入口; 登录按钮状态", "中", "是", ""],
    ["TA-013", "前台商城", "会员登录", "登录成功", "POST /login", "API/会话", "已注册未锁定用户", "loginName、password、verifyCode", "决策表; 等价类", "正确账号密码验证码; 创建 session; 跳转首页/个人中心", "高", "是", "涉及 Session 状态"],
    ["TA-014", "前台商城", "会员登录", "登录失败输入组合", "POST /login", "API", "游客", "空用户名/密码/验证码、错误验证码、错误密码", "决策表", "字段为空分别提示; 验证码错误; 密码错误; 不存在用户", "高", "是", "适合详细测试设计模块"],
    ["TA-015", "前台商城", "会员登录", "锁定用户禁止登录", "POST /login", "API/权限", "被后台锁定用户", "lockedFlag=1", "状态转换; 决策表", "锁定用户无法登录; 解锁后可登录", "高", "是", "与后台会员锁定联动"],
    ["TA-016", "前台商城", "验证码", "商城验证码生成与校验", "GET /common/mall/kaptcha", "API/会话", "游客", "captcha session", "等价类; 安全测试", "验证码图片返回; 多次刷新旧码失效; 大小写/干扰线识别规则", "中", "部分", "自动化可只测接口返回和 session"],
    ["TA-017", "前台商城", "个人中心", "个人信息展示", "GET /personal", "页面/权限", "已登录用户", "昵称、签名、地址", "权限测试; UI检查", "未登录跳转登录; 已登录显示当前用户; 长昵称截断", "中", "是", "被登录拦截器保护"],
    ["TA-018", "前台商城", "个人中心", "修改个人资料", "POST /personal/updateInfo", "API/业务", "已登录用户", "nickName、address、introduceSign", "等价类; 边界值; 安全测试", "正常修改; 空字段保持原值; 超长文本; XSS字符串清洗", "高", "是", "使用 cleanString，安全性可测"],
    ["TA-019", "前台商城", "个人中心", "收货地址维护入口", "GET /personal/addresses", "页面", "已登录用户", "地址页面", "UI检查; 权限测试", "未登录拦截; 页面是否存在; 地址为空状态", "中", "是", "模板可能缺失/功能较弱，需实测"],
    ["TA-020", "前台商城", "退出登录", "商城退出", "GET /logout", "会话", "已登录用户", "session", "状态转换", "退出后 session 清除; 访问购物车/个人中心被拦截", "中", "是", "前后台都有 logout，注意路径冲突"],
    ["TA-021", "前台商城", "购物车", "购物车页面展示", "GET /shop-cart", "页面/业务", "已登录用户", "购物车商品、数量、总价", "等价类; 边界值", "空购物车; 单商品; 多商品; 总价计算; 商品名过长截断", "高", "是", "核心购买链路"],
    ["TA-022", "前台商城", "购物车", "添加商品到购物车", "POST /shop-cart", "API/业务", "已登录用户", "goodsId、goodsCount", "等价类; 边界值", "有效商品; 不存在商品; 下架商品; count=1/5/6/0/负数", "高", "是", "单商品上限=5，但代码未显式拦截 <=0"],
    ["TA-023", "前台商城", "购物车", "重复添加同一商品", "POST /shop-cart", "API/业务", "已登录用户", "已存在 goodsId", "状态转换; 边界值", "重复添加改为更新数量; 数量超过5失败; 原购物车项不重复", "高", "是", "save 会转 update"],
    ["TA-024", "前台商城", "购物车", "购物车容量上限", "POST /shop-cart", "API/业务", "已登录用户", "购物车商品种类数", "边界值", "已有12种再加1成功; 已有13种再加1失败", "高", "是", "Constants.SHOPPING_CART_ITEM_TOTAL_NUMBER=13"],
    ["TA-025", "前台商城", "购物车", "修改购物车商品数量", "PUT /shop-cart", "API/业务", "已登录用户", "cartItemId、goodsCount", "边界值; 权限测试", "数量1/5/6/0; 不存在cartItemId; 修改他人cartItemId", "高", "是", "含 NO_PERMISSION 分支"],
    ["TA-026", "前台商城", "购物车", "删除购物车商品", "DELETE /shop-cart/{cartItemId}", "API/业务", "已登录用户", "cartItemId", "等价类; 权限测试", "删除自己的项; 删除不存在项; 删除他人项; 删除后总价更新", "高", "是", ""],
    ["TA-027", "前台商城", "订单结算", "结算页展示", "GET /shop-cart/settle", "页面/业务", "已登录且购物车非空", "购物车项、总价", "状态转换; 边界值", "空购物车回购物车; 有商品显示总价; 总价异常处理", "高", "是", "生成订单前置"],
    ["TA-028", "前台商城", "订单生成", "保存订单", "GET /saveOrder", "业务流程", "已登录、购物车非空、有地址", "地址、库存、商品状态、价格", "决策表; 状态转换", "无地址; 空购物车; 下架商品; 库存不足; 正常生成订单并扣库存清购物车", "高", "是", "核心交易链路，建议重点模块"],
    ["TA-029", "前台商城", "订单列表", "我的订单列表", "GET /orders?page=...", "页面/API", "已登录用户", "page、订单状态、订单项", "边界值; 权限测试", "第一页; 下一页; 空订单; 只显示当前用户订单", "高", "是", "每页3条"],
    ["TA-030", "前台商城", "订单详情", "查看订单详情", "GET /orders/{orderNo}", "页面/API", "订单所属用户", "orderNo", "等价类; 权限测试", "查看自己的订单; 不存在订单; 查看他人订单禁止", "高", "是", "NO_PERMISSION 分支"],
    ["TA-031", "前台商城", "订单取消", "用户取消订单", "PUT /orders/{orderNo}/cancel", "API/状态", "订单所属用户", "订单状态", "状态转换; 决策表", "待支付可取消; 已完成/已关闭不可取消; 取消后恢复库存", "高", "是", "涉及库存回滚"],
    ["TA-032", "前台商城", "确认收货", "用户确认收货", "PUT /orders/{orderNo}/finish", "API/状态", "订单所属用户", "订单状态", "状态转换", "仅已出库可确认; 非出库状态失败; 他人订单禁止", "高", "是", "订单状态机"],
    ["TA-033", "前台商城", "选择支付方式", "支付方式选择页", "GET /selectPayType?orderNo=...", "页面/权限", "订单所属用户且待支付", "orderNo", "状态转换; 权限测试", "待支付订单展示; 非待支付失败; 他人订单失败", "高", "是", ""],
    ["TA-034", "前台商城", "支付页面", "支付宝/微信支付页", "GET /payPage?orderNo=...&payType=1/2", "页面/业务", "订单所属用户且待支付", "payType", "等价类; 决策表", "payType=1支付宝; payType=2微信; 其它值默认微信; 非待支付失败", "高", "是", "支付类型校验较弱，可作为风险点"],
    ["TA-035", "前台商城", "支付成功回调", "支付成功处理", "GET /paySuccess?orderNo=...&payType=...", "API/状态", "待支付订单", "orderNo、payType", "状态转换; 等价类", "待支付->已支付; 重复回调; 不存在订单; 非法payType", "高", "是", "未校验用户，适合安全风险记录"],
    ["TA-036", "后台管理", "后台登录", "后台登录页", "GET /admin/login", "页面", "管理员", "验证码、账号密码", "UI检查", "页面可访问; 验证码加载; 错误消息显示", "中", "是", "默认 admin/123456"],
    ["TA-037", "后台管理", "后台登录", "管理员登录", "POST /admin/login", "页面表单/会话", "管理员", "userName、password、verifyCode", "决策表", "正确账号密码验证码; 空字段; 错误验证码; 错误密码", "高", "是", "成功后进入 /admin/index"],
    ["TA-038", "后台管理", "后台权限", "后台登录拦截", "/admin/**", "权限/会话", "未登录/已登录管理员", "session loginUser", "权限测试", "未登录访问后台页面跳 /admin/login; 静态资源不拦截", "高", "是", "AdminLoginInterceptor"],
    ["TA-039", "后台管理", "仪表盘", "后台首页", "GET /admin/index", "页面", "管理员", "统计数据/菜单", "UI检查; 权限测试", "菜单显示; 路由跳转; session 昵称显示", "中", "是", ""],
    ["TA-040", "后台管理", "管理员资料", "查看管理员资料", "GET /admin/profile", "页面", "管理员", "loginUserId", "权限测试; UI检查", "正常显示用户名昵称; session失效跳登录", "中", "是", ""],
    ["TA-041", "后台管理", "管理员资料", "修改管理员昵称/用户名", "POST /admin/profile/name", "API/表单", "管理员", "loginUserName、nickName", "等价类; 边界值", "正常修改; 空字段; 超长字段; 特殊字符", "中", "是", ""],
    ["TA-042", "后台管理", "管理员资料", "修改管理员密码", "POST /admin/profile/password", "API/状态", "管理员", "originalPassword、newPassword", "决策表; 边界值", "原密码正确; 原密码错误; 空字段; 修改成功后 session 清除需重新登录", "高", "是", "认证安全相关"],
    ["TA-043", "后台管理", "轮播图管理", "轮播图列表", "GET /admin/carousels/list", "API/分页", "管理员", "page、limit", "边界值", "缺少分页参数; page边界; 排序是否正确", "中", "是", ""],
    ["TA-044", "后台管理", "轮播图管理", "新增轮播图", "POST /admin/carousels/save", "API/CRUD", "管理员", "carouselUrl、redirectUrl、rank", "等价类; 边界值", "必填为空; URL非法; rank边界; 新增后前台首页显示", "中", "是", "前后台联动"],
    ["TA-045", "后台管理", "轮播图管理", "更新/删除轮播图", "POST update/delete", "API/CRUD", "管理员", "carouselId、ids", "等价类; 边界值", "更新不存在ID; 批量删除空数组; 删除后前台不显示", "中", "是", ""],
    ["TA-046", "后台管理", "分类管理", "分类列表与三级联动", "GET /admin/categories; listForSelect", "页面/API", "管理员", "categoryLevel、parentId、backParentId", "等价类; 边界值", "一级/二级/三级分类; 无子分类; parentId非法; 返回级联列表正确", "高", "是", "商品编辑依赖分类"],
    ["TA-047", "后台管理", "分类管理", "新增分类", "POST /admin/categories/save", "API/CRUD", "管理员", "categoryName、categoryLevel、parentId、rank", "等价类; 边界值", "同级同名重复; 空名称; 等级非法; parentId不存在", "高", "是", "影响搜索和商品归类"],
    ["TA-048", "后台管理", "分类管理", "修改/删除分类", "POST update/delete", "API/CRUD", "管理员", "categoryId、ids", "等价类; 状态影响", "修改名称排序; 删除有子分类/有商品分类; 删除后前台隐藏", "高", "是", "数据一致性重点"],
    ["TA-049", "后台管理", "商品管理", "商品列表查询", "GET /admin/goods/list", "API/分页", "管理员", "page、limit、goodsName、goodsSellStatus", "边界值; 等价类", "缺少分页; 按名称搜索; 上架/下架过滤; 页码越界", "中", "是", ""],
    ["TA-050", "后台管理", "商品管理", "新增商品", "POST /admin/goods/save", "API/CRUD", "管理员", "名称、简介、价格、分类、库存、图片、详情", "等价类; 边界值; 决策表", "必填为空; 价格0/负数/超大; 库存0/负数; 分类非三级; 重复商品", "高", "是", "商品数据是购物链路源头"],
    ["TA-051", "后台管理", "商品管理", "编辑商品", "POST /admin/goods/update", "API/CRUD", "管理员", "goodsId及商品字段", "等价类; 边界值", "不存在ID; 必填为空; 修改价格库存状态; 修改后前台详情一致", "高", "是", ""],
    ["TA-052", "后台管理", "商品管理", "上架/下架商品", "PUT /admin/goods/status/{sellStatus}", "API/状态", "管理员", "ids、sellStatus=0/1", "状态转换; 边界值", "空ids; 非法状态; 上架后可搜索/购买; 下架后不可下单", "高", "是", "与订单生成联动"],
    ["TA-053", "后台管理", "商品管理", "商品图片/富文本上传", "POST /admin/upload/file(s)", "API/文件", "管理员", "MultipartFile", "等价类; 安全测试", "图片上传; 多文件; 空文件; 非图片; 超大文件; 路径穿越文件名", "高", "部分", "文件上传安全风险高"],
    ["TA-054", "后台管理", "首页配置", "首页配置列表", "GET /admin/indexConfigs/list", "API/分页", "管理员", "configType、page、limit", "等价类; 边界值", "不同配置类型; 缺少分页; 越界页", "中", "是", "新品/热门/推荐"],
    ["TA-055", "后台管理", "首页配置", "新增/修改首页配置", "POST save/update", "API/CRUD", "管理员", "configName、configType、goodsId、rank", "等价类; 边界值", "重复配置; 商品不存在; rank排序; 保存后首页对应区域显示", "中", "是", ""],
    ["TA-056", "后台管理", "首页配置", "删除首页配置", "POST /admin/indexConfigs/delete", "API/CRUD", "管理员", "ids", "边界值; 数据一致性", "空ids; 不存在ID; 删除后首页不显示", "中", "是", ""],
    ["TA-057", "后台管理", "订单管理", "后台订单列表", "GET /admin/orders/list", "API/分页", "管理员", "page、limit、orderNo、orderStatus", "边界值; 等价类", "分页; 状态过滤; 订单号搜索; 无结果", "高", "是", "运营核心功能"],
    ["TA-058", "后台管理", "订单管理", "修改订单金额/地址", "POST /admin/orders/update", "API/状态", "管理员", "orderId、totalPrice、userAddress", "状态转换; 边界值", "待支付/已支付可改; 已出库后不可改; 金额<=0; 地址为空", "高", "是", "影响交易金额"],
    ["TA-059", "后台管理", "订单管理", "查看订单明细", "GET /admin/order-items/{id}", "API", "管理员", "orderId", "等价类", "有效订单; 无订单项; 不存在订单", "中", "是", ""],
    ["TA-060", "后台管理", "订单管理", "配货完成", "POST /admin/orders/checkDone", "API/状态", "管理员", "ids、订单状态", "状态转换; 决策表", "仅已支付订单可配货; 多选混合状态; 空ids", "高", "是", "订单状态机"],
    ["TA-061", "后台管理", "订单管理", "出库", "POST /admin/orders/checkOut", "API/状态", "管理员", "ids、订单状态", "状态转换; 决策表", "已支付/已配货可出库; 其他状态失败; 多选混合状态", "高", "是", ""],
    ["TA-062", "后台管理", "订单管理", "关闭订单", "POST /admin/orders/close", "API/状态", "管理员", "ids、订单状态", "状态转换; 决策表", "待支付/已支付可关闭; 已完成/已关闭不可关闭; 关闭后恢复库存", "高", "是", "库存一致性风险"],
    ["TA-063", "后台管理", "会员管理", "会员列表", "GET /admin/users/list", "API/分页", "管理员", "page、limit、loginName、lockedFlag", "边界值; 等价类", "分页; 搜索; 锁定状态过滤; 缺少分页参数", "中", "是", ""],
    ["TA-064", "后台管理", "会员管理", "锁定/解锁会员", "POST /admin/users/lock/{lockStatus}", "API/状态", "管理员", "ids、lockStatus", "状态转换; 边界值", "lockStatus=0/1; 非法状态; 空ids; 锁定后用户不能登录", "高", "是", "与前台登录联动"],
    ["TA-065", "公共/异常", "错误页", "400/404/5xx 错误处理", "/error, invalid URL", "页面/异常", "任意用户", "HTTP status", "异常测试", "不存在路由; 参数类型错误; 服务异常; 错误页文案和状态码", "中", "是", "ErrorPageController"],
    ["TA-066", "公共/安全", "认证会话隔离", "前台/后台 session", "安全/权限", "游客、会员、管理员", "Session Cookie", "权限测试; 安全测试", "会员不能访问后台; 管理员登录不等于会员登录; 退出互不污染", "高", "是", "两个登录体系"],
    ["TA-067", "公共/安全", "CSRF/未授权状态修改", "所有 POST/PUT/DELETE", "安全/API", "未登录或跨站请求", "Cookie、Referer、CSRF token", "安全测试", "未登录是否拦截; 是否存在CSRF防护; 直接调用状态修改接口", "高", "部分", "项目未见显式CSRF token"],
    ["TA-068", "公共/性能", "核心页面响应时间", "首页/搜索/详情/购物车/订单", "性能", "正常数据量", "响应时间、并发用户", "性能测试; 边界值", "单用户响应; 10/50并发; 大量商品分页查询", "中", "是", "可用于 NFR"],
    ["TA-069", "公共/兼容性", "浏览器与响应式显示", "前后台页面", "UI/NFR", "Chrome/Edge/移动宽度", "布局、图片、表格", "兼容性测试; UI检查", "桌面/移动宽度; 表格溢出; 图片缺失; 中文乱码", "中", "部分", "README 有乱码但页面可能正常"],
    ["TA-070", "公共/数据一致性", "库存与订单事务一致性", "下单/取消/关闭订单", "业务/数据库", "有库存商品", "库存、购物车、订单项", "状态转换; 事务测试", "下单扣库存; 下单失败回滚; 取消/关闭恢复库存; 并发下单超卖", "高", "部分", "最适合深入白盒+黑盒结合"],
]

ROUTE_HEADERS = ["区域", "方法", "路径", "控制器/功能", "认证要求"]
ROUTES = [
    ["前台", "GET", "/, /index, /index.html", "首页", "游客"],
    ["前台", "GET", "/search, /search.html", "商品搜索", "游客"],
    ["前台", "GET", "/goods/detail/{goodsId}", "商品详情", "游客"],
    ["前台", "GET", "/login", "会员登录页", "游客"],
    ["前台", "POST", "/login", "会员登录", "游客+验证码"],
    ["前台", "GET", "/register", "会员注册页", "游客"],
    ["前台", "POST", "/register", "会员注册", "游客+验证码"],
    ["前台", "GET", "/personal", "个人中心", "会员"],
    ["前台", "POST", "/personal/updateInfo", "修改个人资料", "会员"],
    ["前台", "GET", "/logout", "会员退出", "会员"],
    ["前台", "GET", "/shop-cart", "购物车页", "会员"],
    ["前台", "POST", "/shop-cart", "添加购物车", "会员"],
    ["前台", "PUT", "/shop-cart", "修改购物车", "会员"],
    ["前台", "DELETE", "/shop-cart/{id}", "删除购物车项", "会员"],
    ["前台", "GET", "/shop-cart/settle", "结算页", "会员"],
    ["前台", "GET", "/saveOrder", "生成订单", "会员"],
    ["前台", "GET", "/orders", "我的订单", "会员"],
    ["前台", "GET", "/orders/{orderNo}", "订单详情", "会员"],
    ["前台", "PUT", "/orders/{orderNo}/cancel", "取消订单", "会员"],
    ["前台", "PUT", "/orders/{orderNo}/finish", "确认收货", "会员"],
    ["前台", "GET", "/selectPayType", "选择支付方式", "会员"],
    ["前台", "GET", "/payPage", "支付页面", "会员"],
    ["前台", "GET", "/paySuccess", "支付成功处理", "会员/接口"],
    ["公共", "GET", "/common/mall/kaptcha", "商城验证码", "游客"],
    ["公共", "GET", "/common/kaptcha", "后台验证码", "游客"],
    ["后台", "GET", "/admin/login", "后台登录页", "游客"],
    ["后台", "POST", "/admin/login", "管理员登录", "管理员+验证码"],
    ["后台", "GET", "/admin/index", "后台首页", "管理员"],
    ["后台", "GET", "/admin/profile", "管理员资料", "管理员"],
    ["后台", "POST", "/admin/profile/password", "修改密码", "管理员"],
    ["后台", "POST", "/admin/profile/name", "修改名称", "管理员"],
    ["后台", "GET", "/admin/logout", "后台退出", "管理员"],
    ["后台", "GET", "/admin/carousels", "轮播图页", "管理员"],
    ["后台", "GET", "/admin/carousels/list", "轮播图列表", "管理员"],
    ["后台", "POST", "/admin/carousels/save", "新增轮播图", "管理员"],
    ["后台", "POST", "/admin/carousels/update", "更新轮播图", "管理员"],
    ["后台", "GET", "/admin/carousels/info/{id}", "轮播图详情", "管理员"],
    ["后台", "POST", "/admin/carousels/delete", "删除轮播图", "管理员"],
    ["后台", "GET", "/admin/categories", "分类页", "管理员"],
    ["后台", "GET", "/admin/categories/list", "分类列表", "管理员"],
    ["后台", "GET", "/admin/categories/listForSelect", "分类级联", "管理员"],
    ["后台", "POST", "/admin/categories/save", "新增分类", "管理员"],
    ["后台", "POST", "/admin/categories/update", "更新分类", "管理员"],
    ["后台", "GET", "/admin/categories/info/{id}", "分类详情", "管理员"],
    ["后台", "POST", "/admin/categories/delete", "删除分类", "管理员"],
    ["后台", "GET", "/admin/goods", "商品管理页", "管理员"],
    ["后台", "GET", "/admin/goods/edit", "新增商品页", "管理员"],
    ["后台", "GET", "/admin/goods/edit/{goodsId}", "编辑商品页", "管理员"],
    ["后台", "GET", "/admin/goods/list", "商品列表", "管理员"],
    ["后台", "POST", "/admin/goods/save", "新增商品", "管理员"],
    ["后台", "POST", "/admin/goods/update", "更新商品", "管理员"],
    ["后台", "GET", "/admin/goods/info/{id}", "商品详情API", "管理员"],
    ["后台", "PUT", "/admin/goods/status/{sellStatus}", "上下架商品", "管理员"],
    ["后台", "GET", "/admin/indexConfigs", "首页配置页", "管理员"],
    ["后台", "GET", "/admin/indexConfigs/list", "首页配置列表", "管理员"],
    ["后台", "POST", "/admin/indexConfigs/save", "新增首页配置", "管理员"],
    ["后台", "POST", "/admin/indexConfigs/update", "更新首页配置", "管理员"],
    ["后台", "GET", "/admin/indexConfigs/info/{id}", "首页配置详情", "管理员"],
    ["后台", "POST", "/admin/indexConfigs/delete", "删除首页配置", "管理员"],
    ["后台", "GET", "/admin/orders", "订单管理页", "管理员"],
    ["后台", "GET", "/admin/orders/list", "订单列表", "管理员"],
    ["后台", "POST", "/admin/orders/update", "更新订单", "管理员"],
    ["后台", "GET", "/admin/order-items/{id}", "订单项", "管理员"],
    ["后台", "POST", "/admin/orders/checkDone", "配货完成", "管理员"],
    ["后台", "POST", "/admin/orders/checkOut", "出库", "管理员"],
    ["后台", "POST", "/admin/orders/close", "关闭订单", "管理员"],
    ["后台", "GET", "/admin/users", "会员管理页", "管理员"],
    ["后台", "GET", "/admin/users/list", "会员列表", "管理员"],
    ["后台", "POST", "/admin/users/lock/{lockStatus}", "锁定/解锁会员", "管理员"],
    ["后台", "POST", "/admin/upload/file", "单文件上传", "管理员"],
    ["后台", "POST", "/admin/upload/files", "多文件上传", "管理员"],
]

STARTUP_ROWS = [
    ["项目", "newbee-mall Spring Boot + Thymeleaf + MyBatis + MySQL"],
    ["本地访问地址", "http://localhost:28089"],
    ["数据库", "MySQL: localhost:3306/newbee_mall_db, root/123456"],
    ["本次处理", r"已创建临时 MySQL 数据目录 E:\College\3down\ST\FINAL\tmp\mysql8，并导入 src/main/resources/newbee_mall_schema.sql"],
    ["后台默认账号", "admin / 123456"],
    ["前台会员", "可通过 /register 注册；也可从数据库 tb_newbee_mall_user 选择已有用户"],
    ["推荐重点测试模块", "会员登录注册、购物车、下单支付、订单状态流转、后台商品管理、后台订单管理、文件上传安全"],
    ["建议自动化框架", "Playwright 或 Selenium 做端到端；JUnit/SpringBootTest 做 Service/Controller；也可用 PyTest+requests 做接口"],
]


def style_sheet(ws):
    ws.freeze_panes = "A2"
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in ws.columns:
        width = 10
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, 36))
        ws.column_dimensions[get_column_letter(col[0].column)].width = width


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Testable Areas"
    ws.append(HEADERS)
    for row in ROWS:
        ws.append(row)
    ws.auto_filter.ref = f"A1:M{ws.max_row}"

    routes = wb.create_sheet("Routes")
    routes.append(ROUTE_HEADERS)
    for row in ROUTES:
        routes.append(row)
    routes.auto_filter.ref = f"A1:E{routes.max_row}"

    startup = wb.create_sheet("Startup")
    for row in STARTUP_ROWS:
        startup.append(row)

    for sheet in wb.worksheets:
        style_sheet(sheet)

    priority_col = 11
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, priority_col)
        if cell.value == "高":
            cell.fill = PatternFill("solid", fgColor="F4CCCC")
        elif cell.value == "中":
            cell.fill = PatternFill("solid", fgColor="FFF2CC")
        else:
            cell.fill = PatternFill("solid", fgColor="D9EAD3")

    wb.save(OUT)
    print(OUT)
    print(f"testable_rows={len(ROWS)} routes={len(ROUTES)}")


if __name__ == "__main__":
    main()
